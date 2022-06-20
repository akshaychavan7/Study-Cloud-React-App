import openpyxl
import json
  
# load excel with its path
wrkbk = openpyxl.load_workbook("Study Cloud Questions.xlsx")
  
sh = wrkbk.active
questions = []

def getAnswerNumericValue(charAnswer):
    match charAnswer:
        case "A": return 1
        case "B": return 2
        case "C": return 3
        case "D": return 4

# iterate through excel and display data
for i in range(5, sh.max_row+1):
    print("\n")
    print("Row ", i, " data :")

    questionObj = {
            "question": sh.cell(row=i, column=1).value,
            "options": [
                        sh.cell(row=i, column=2).value,
                        sh.cell(row=i, column=3).value,
                        sh.cell(row=i, column=4).value,
                        sh.cell(row=i, column=5).value
                        ],
            "correctAnswer": [getAnswerNumericValue(sh.cell(row=i, column=6).value)],
            "explaination": "",
    "explainationLink": "",
    "image": ""
        }
    questions.append(questionObj)

print(json.dumps(questions))
    

        
